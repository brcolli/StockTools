import random
import pandas as pd
import openpyxl
from openpyxl.styles import *
from openpyxl.chart import AreaChart, Reference, Series
import os
import numpy as np
import string

# Todos in priority order
# TODO: Create a database and indexing system for storing all ending capitals vs a single cell string in Excel
# TODO: Export the input data for simulations into Excel and database so it can be reviewed later
# TODO: Add Histogram options for other sets of data (Sortino Ratio, Alpha, etc.)
# TODO: Fix Histogram style and labels getting deleted after reading Excel file


class KellyModel:
    """
    Used to simulate trading with the kelly model, either single simulation of n trades or multiple simulations of n
    trades.
    """

    def __init__(self, excel_path: str, beg_capital=10000, edge=0.5, edge_range=0, glr=1.15, fractional_kelly=0.25,
                 l_ceiling=0.2, w_ceiling=0.5, num_runs=100, tax_rate=0.1, snp=0.1478, expected_return=0.0008):
        """
        :param excel_path: Filepath to the excel file to be operated on
        :type excel_path: str

        :param beg_capital: Starting capital (USD)
        :type beg_capital: float
        :param edge: Average chance of winning a trade in (0, 1)
        :type edge: float
        :param  edge_range: The range for the edge for any specific trade:
                            (Current Edge = Edge + random.uniform(-range, +range)
        :type edge_range: float
        :param glr: Gain loss ratio in (0, 1)
        :type glr: float
        :param fractional_kelly: The fraction of the Kelly model to use in (0, 1)
        :type fractional_kelly: float
        :param l_ceiling: Loss ceiling percentage wise for any given trade in (0, 1)
        :type l_ceiling: float
        :param w_ceiling: Winn ceiling percentage wise for any give trade in (0, 1)
        :type w_ceiling: float
        :param num_runs: Number of trades to do perform in a single simulation
        :type num_runs: int
        :param tax_rate: The tax rate (used to calculate alpha) percentage wise in (0, 1)
        :type tax_rate: float
        :param snp: SNP 500 returns (used to calculate alpha) percentage wise in (0, 1)
        :type snp: float
        :param expected_return: Expected return per trade (used to calculate Sortino Ratio) percentage wise in (0, 1)
        :type expected_return: float
        """

        self.beg_capital, self.edge, self.edge_range, self.glr, self.fractional_kelly, self.l_ceiling, self.w_ceiling, \
        self.num_runs = \
            beg_capital, edge, edge_range, glr, fractional_kelly, l_ceiling, w_ceiling, num_runs

        self.tax_rate = tax_rate
        self.snp = snp
        self.expected_return = expected_return

        self.curr_capital = self.beg_capital
        self.num_plays = 0
        self.num_bets = 0
        self.num_wins = 0
        self.num_losses = 0
        self.largest_gain = 0
        self.largest_loss = 0
        self.largest_loss_streak = 0
        self.largest_win_streak = 0
        self.curr_streak = 0

        self.average_gain = 0
        self.average_loss = 0
        self.excess_return = 0
        self.negative_excess_return = 0
        self.square_negative_excess_return = 0

        self.win_v_loss = 0
        self.gross_return = 0
        self.net_return = 0
        self.alpha = 0
        self.win_rate = 0
        self.multiple_on_invested_capital = 0
        self.expected_gl = 0
        self.actual_gl = 0
        self.downside_risk = 0
        self.average_excess_return = 0
        self.sortino = 0

        self.ending_capital = [self.curr_capital]

        self.keys = ['Ending Capital List', 'Final Ending Capital', 'Wins vs Losses', 'Largest Winning Streak',
                     'Largest Losing Streak',
                     'Gross Return', 'Tax Rate', 'Net Return', 'YTD S&P 500 Return', 'Alpha', 'W', 'L',
                     'Total Trades', 'Win Rate', 'Multiple on Invested Capital', 'Largest Gain', 'Largest Loss',
                     'Average Gain', 'Average Loss', 'Expected Gain/(Loss) Per Trade', 'Actual Gain/Loss Ratio',
                     'Downside Risk', 'Average Excess Return', 'Sortino Ratio']

        self.excel_path = excel_path
        if not os.path.isfile(excel_path):
            wb = openpyxl.Workbook()
            wb.save(excel_path)
            wb.close()

    def single_trade(self):
        """
        Simulates a single trade and records the results
        """

        # Edge calculation for this specific trade using edge range
        cur_edge = self.edge + random.uniform(-1 * self.edge_range, self.edge_range)

        # Kelly model bet size
        bet = (cur_edge - ((1 - cur_edge) / self.glr)) * self.fractional_kelly

        trade_return = 0
        self.num_plays += 1
        if bet > 0:
            self.num_bets += 1

            to_trade = bet * self.curr_capital

            # Win
            if random.random() <= cur_edge:
                self.num_wins += 1
                if self.curr_streak < 0:
                    self.curr_streak = 1
                else:
                    self.curr_streak += 1

                if self.curr_streak > self.largest_win_streak:
                    self.largest_win_streak = self.curr_streak

                percent_gain_realized = random.uniform(0, self.w_ceiling)
                trade_return = percent_gain_realized * 1 * to_trade
                diff_in_returns = percent_gain_realized - self.expected_return
                self.excess_return += diff_in_returns
                if diff_in_returns < 0:
                    self.negative_excess_return += diff_in_returns
                    self.square_negative_excess_return += ((-1 * diff_in_returns) ** 2)

                self.average_gain += trade_return
                if trade_return > self.largest_gain:
                    self.largest_gain = trade_return

            # Loss
            else:
                self.num_losses += 1
                if self.curr_streak > 0:
                    self.curr_streak = -1
                else:
                    self.curr_streak -= 1

                if self.curr_streak < self.largest_loss_streak:
                    self.largest_loss_streak = self.curr_streak

                percent_loss_realized = -1 * random.uniform(0, self.l_ceiling)
                trade_return = percent_loss_realized * to_trade
                diff_in_returns = percent_loss_realized - self.expected_return
                self.excess_return += diff_in_returns
                if diff_in_returns < 0:
                    self.negative_excess_return += diff_in_returns
                    self.square_negative_excess_return += ((-1 * diff_in_returns) ** 2)

                self.average_loss -= trade_return
                if (-1 * trade_return) > self.largest_loss:
                    self.largest_loss = (-1 * trade_return)

            self.curr_capital += trade_return
            self.ending_capital.append(self.curr_capital)

    def all_trades(self):
        """
        Simulates a trade (num_runs) times with each trade compounding on the previous
        """
        for i in range(self.num_runs):
            self.single_trade()

    def calc_ending_scores(self):
        """
        Calculates ending variables after all the trades are completed for a single simulation
        """
        self.win_v_loss = self.num_wins - self.num_losses
        self.gross_return = (self.curr_capital - self.beg_capital) / self.beg_capital
        self.net_return = self.gross_return * (1 - self.tax_rate)
        self.alpha = self.net_return - self.snp
        self.win_rate = self.num_wins / self.num_bets
        self.multiple_on_invested_capital = self.curr_capital / self.beg_capital
        self.average_gain = (self.average_gain / self.num_wins)
        self.average_loss = (self.average_loss / self.num_losses)
        self.excess_return = self.excess_return / self.num_bets
        self.negative_excess_return = self.negative_excess_return / self.num_bets
        self.square_negative_excess_return = self.square_negative_excess_return / self.num_bets
        self.expected_gl = (self.average_gain * self.edge) - (self.average_loss * (1 - self.edge))
        self.actual_gl = self.average_gain / self.average_loss
        self.downside_risk = (self.square_negative_excess_return ** 0.5)
        self.average_excess_return = self.excess_return
        self.sortino = self.average_excess_return / self.downside_risk

    def to_dict(self):
        """
        Exports values of a single simulation to a dictionary

        :return: dictionary of result values
        :rtype: dict
        """
        values = [self.ending_capital, self.curr_capital, self.win_v_loss, self.largest_win_streak,
                  self.largest_loss_streak,
                  self.gross_return, self.tax_rate, self.net_return, self.snp, self.alpha, self.num_wins,
                  self.num_losses, self.num_bets, self.win_rate, self.multiple_on_invested_capital, self.largest_gain,
                  self.largest_loss, self.average_gain, self.average_loss, self.expected_gl, self.actual_gl,
                  self.downside_risk, self.average_excess_return, self.sortino]

        return dict(zip(self.keys, values))

    def single_sim(self):
        """
        Wrapper for running a single simulation

        :return: Pandas dataframe with results
        :rtype: pd.DataFrame
        """
        self.all_trades()
        self.calc_ending_scores()
        return self.to_dict()

    def reset_sim(self):
        """
        Resets the sim to starting conditions (used in the case multiple iterations are being run)
        """
        self.curr_capital = self.beg_capital
        self.num_plays = 0
        self.num_bets = 0
        self.num_wins = 0
        self.num_losses = 0
        self.largest_gain = 0
        self.largest_loss = 0
        self.largest_loss_streak = 0
        self.largest_win_streak = 0
        self.curr_streak = 0

        self.average_gain = 0
        self.average_loss = 0
        self.excess_return = 0
        self.negative_excess_return = 0
        self.square_negative_excess_return = 0

        self.win_v_loss = 0
        self.gross_return = 0
        self.net_return = 0
        self.alpha = 0
        self.win_rate = 0
        self.multiple_on_invested_capital = 0
        self.expected_gl = 0
        self.actual_gl = 0
        self.downside_risk = 0
        self.average_excess_return = 0
        self.sortino = 0
        self.ending_capital = [self.curr_capital]

    def multi_sim(self, num):
        """
        Wrapper for running multiple simulations (ie running multiple sets of 100 trades to see what outcomes are
        possible)

        :param num: Number of simulation to run (ie how many times to simulate 100 trades in a row)\
        :type num: int

        :return: Pandas dataframe with results
        :rtype: pd.DataFrame
        """
        df = pd.DataFrame()

        for i in range(num):
            self.reset_sim()
            df = df.append(self.single_sim(), ignore_index=True)
            df = df[self.keys]

        return df

    @staticmethod
    def save_df(df: pd.DataFrame, name: str):
        """
        Save a dataframe to a csv which can then be opened in Excel for analysis

        :param df: dataframe to save
        :type df: pd.DataFrame
        :param name: file path and name
        :type name: str
        """
        df.to_csv(name, index=False)

    def to_excel(self, df: pd.DataFrame, sheet_name: str):
        """
        Exports a results dataframe to an excel sheet, workbook specified by self.excel_path

        :param df: Results dataframe with keys as in self.keys
        :type df: pd.DataFrame
        :param sheet_name: Name of sheet to export to
        :type sheet_name: str
        """

        with pd.ExcelWriter(self.excel_path, mode='a') as writer:
            df.to_excel(writer, sheet_name=sheet_name)

        wb = openpyxl.load_workbook(filename=self.excel_path)
        self.histogram(df['Final Ending Capital'], sheet_name, wb)
        self.format_sheet(sheet_name, wb)
        wb.save(filename=self.excel_path)

    def multi_sim_to_excel(self, num: int, sheet_name: str):
        """
        Combines the methods multi_sim and to_excel for convenience

        :param num: Number of simulations to run
        :type num: int
        :param sheet_name: Name of Excel sheet
        :type sheet_name: str

         :return: Dataframe of results
         :rtype: pd.DataFrame
        """
        df = self.multi_sim(num)
        self.to_excel(df, sheet_name)
        return df

    @staticmethod
    def column_num_from_letter(name: str):
        """
        Converts a column name like "AA" or "AB" or "X" to a number

        :param name: Column name
        :type name: str

        :return: Numerical value
        :rtype: int
        """
        name = name.lower()
        num = 0
        for i in range(len(name)):
            p = len(name) - i - 1
            num += (26 ** p) * (string.ascii_lowercase.index(name[i]) + 1)

        return num

    def int_to_column_id(self, num):
        """
        Converts any positive integer to Base26(letters only) with no 0th
        case. Useful for applications such as spreadsheet columns to determine which
        Letterset goes with a positive integer.

        :param num: Number to convert to column name
        :type num: int

        :return: Column name
        :rtype: str
        """

        alphabet = string.ascii_uppercase
        n = len(alphabet)
        if num < 0:
            raise ValueError("Input should be a non-negative integer.")
        elif num == 0:
            return ""
        else:
            q, r = divmod(num - 1, n)
            return self.int_to_column_id(q) + alphabet[r]

    def histogram(self, column: list[float], sheet_name: str, wb: openpyxl.Workbook()):
        """
        Generates and plots an area style histogram in Excel of any dataset (though is intended for the Final Ending
        Capital column of the results dataframe).

        :param column: The column of data to make a histogram from
        :type column: list[float]
        :param sheet_name: Name of the sheet to plot the histogram in
        :type sheet_name: str
        :param wb: An instance of the workbook class from openpyxl
        :type wb: openpyxl.Workbook()
        """
        num_bins = 20  # Sets the number of bins, useful to keep # bins the same for comparing histograms
        hist, bins = np.histogram(column, bins=num_bins)  # hist, bins
        ws = wb[sheet_name]
        columns = ['AA', 'AB']  # Cell columns to plop the bins and frequencies in
        data = [bins[1:], hist]
        for c, d in zip(columns, data):
            for i in range(len(d)):
                ws[c + str(i + 2)] = d[i]

        ws[columns[0] + '1'] = 'Bins'
        ws[columns[1] + '1'] = 'Frequencies'

        chart = AreaChart()
        chart.title = "Ending Capital Distribution"
        chart.style = 13
        chart.x_axis.title = 'Ending Capital'
        chart.y_axis.title = 'Frequency'
        AA, AB = [self.column_num_from_letter(c) for c in columns]
        cats = Reference(ws, min_col=AA, min_row=2, max_row=num_bins + 2)
        data = Reference(ws, min_col=AB, min_row=2, max_row=num_bins + 2)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, str(columns[0]) + str(num_bins + 5))  # Second parameter is where the chart is placed,
                                                                    # can change to 'A1' if more convenient

    @staticmethod
    def as_text(value: any):
        """
        Gets the string of a value with edge case None

        :param value: Value to turn into string
        :type value: Any

        :return: String of value
        :rtype: str
        """
        if value is None:
            return ""
        return str(value)

    def format_sheet(self, sheet_name: str, wb: openpyxl.Workbook()):
        """
        Formats an excel sheet including number types and column widths

        :param sheet_name: Name of the sheet to format
        :type sheet_name: str

        :param wb: An instance of the workbook class from openpyxl
        :type wb: openpyxl.Workbook()
        """

        ws = wb[sheet_name]

        # Each tuple is a list of column letters and a string of their respective numerical format
        # Formats found here: https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html
        columns_formats = [
            (['G', 'H', 'I', 'J', 'K', 'O', 'W', 'X'], '0.00%'),
            (['P', 'U', 'V', 'Y'], '#,##0.00'),
            (['C', 'Q', 'R', 'S', 'T', 'AA'], '"$"#,##0.00_);("$"#,##0.00)')
        ]

        # Applies numerical formats to all cells
        for col, form in columns_formats:
            for c in col:
                for i in range(2, len(list(ws.rows)) + 1):
                    cell = ws[c + str(i)]
                    cell.number_format = form

        # Applies column width to all cells except the first (containing a long string)
        c = 0
        for column_cells in list(ws.columns)[1:]:
            if c == 0:
                c += 1
            else:
                length = max(len(self.as_text(cell.value)) for cell in column_cells) + 2
                ws.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length
